import json
import xlwt
import xlrd
import openpyxl
import ast
from openpyxl.styles import Alignment

class Xls_resolve(object):
    path = None

    def __init__(self, path):
        self.path = path

    def test(self):
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook.worksheets[0]

        for i in range( 3):
            print(i)
            print(sheet.cell(row=i+1, column=2).value)

    def save_column_data(self, column_index:int, column_data:list, start_row_index=1):
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook.worksheets[0]

        row_number = sheet.max_row
        print(f'table max row size:{row_number}')
        print(f'column data size:{len(column_data)}')
        for i in range(len(column_data)):
            sheet.cell(row=start_row_index+i, column=column_index+1).value = column_data[i]
        workbook.save(self.path)

    def append_table_data(self, record_list:list):
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook.worksheets[0]

        max_row_number = sheet.max_row
        print(f'table row size:{max_row_number}')
        for index in range(len(record_list)):
            record = record_list[index]

            sheet.cell(row=max_row_number+index+1, column=1).value = record[0]
            sheet.cell(row=max_row_number+index+1, column=2).value = record[1]
            sheet.cell(row=max_row_number+index+1, column=3).value = record[2]
            sheet.cell(row=max_row_number+index+1, column=4).value = record[3]
            sheet.cell(row=max_row_number+index+1, column=5).value = record[4]
            sheet.cell(row=max_row_number+index+1, column=6).value = record[5]
            sheet.cell(row=max_row_number+index+1, column=7).value = record[6]
            sheet.cell(row=max_row_number+index+1, column=8).value = record[7]
            sheet.cell(row=max_row_number+index+1, column=9).value = record[8]
            sheet.cell(row=max_row_number+index+1, column=10).value = record[9]
            sheet.cell(row=max_row_number+index+1, column=11).value = record[10]
            sheet.cell(row=max_row_number+index+1, column=12).value = record[11]
            sheet.cell(row=max_row_number+index+1, column=13).value = record[12]
            sheet.cell(row=max_row_number+index+1, column=14).value = record[13]
            sheet.cell(row=max_row_number+index+1, column=15).value = record[14]
            sheet.cell(row=max_row_number+index+1, column=16).value = record[15]
            sheet.cell(row=max_row_number+index+1, column=17).value = record[16]
            sheet.cell(row=max_row_number+index+1, column=18).value = record[17]
            sheet.cell(row=max_row_number+index+1, column=19).value = record[18]
            sheet.cell(row=max_row_number+index+1, column=20).value = record[19]
            sheet.cell(row=max_row_number+index+1, column=21).value = record[20]
            sheet.cell(row=max_row_number+index+1, column=22).value = record[21]
            sheet.cell(row=max_row_number+index+1, column=23).value = record[22]
            sheet.cell(row=max_row_number+index+1, column=24).value = record[23]
            sheet.cell(row=max_row_number+index+1, column=25).value = record[24]
            sheet.cell(row=max_row_number+index+1, column=26).value = record[25]
            sheet.cell(row=max_row_number+index+1, column=27).value = record[26]
        workbook.save(self.path)


    #column_index 从0开始
    #start_row_index 从0开始,默认取1即第二行开始
    def read_column_data(self, column_index, start_row_index=1):
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook.worksheets[0]

        #rows = sheet.rows
        cols = sheet.columns
        col_data = []
        for cell in list(cols)[column_index][start_row_index:]:
            col_data.append(cell.value)
        print(f'col_data size:{len(col_data)}')
        return col_data

    def read_cell_data(self, column_index, row_index):
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook.worksheets[0]

        #rows = sheet.rows
        cols = sheet.columns
        cell = list(cols)[column_index][row_index]
        #print(f'cell value:{cell.value}')
        return cell.value

    def write_cell_data(self, column_index, row_index, value):
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook.worksheets[0]
        #cols = sheet.columns
        #print('cell value:')
        #print(sheet.cell(row=row_index+1, column=column_index+1).value)
        #print(f'value:{value}')
        sheet.cell(row=row_index+1, column=column_index+1).value = value
        workbook.save(self.path)


    def get_left_result(self, id_list:list):
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook.worksheets[0]
        result_value_list = []
        rows = list(sheet.rows)
        for index in range(len(id_list)):
            id = id_list[index]
            row = rows[int(id)]
            row_result = []
            row_result.append(row[10].value)
            row_result.append(row[11].value)
            row_result.append(row[12].value)
            row_result.append(row[13].value)
            row_result.append(row[14].value)
            row_result.append(row[15].value)
            result_value_list.append(row_result)
        #print(result_value_list)
        return result_value_list


    def write_right_result(self,id_list:list,result_list:list):
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook.worksheets[0]
        rows = list(sheet.rows)
        for index in range(len(id_list)):
            id = id_list[index]
            row = rows[int(id)]
            row_result = result_list[index]
            row[10].value = row_result[0]
            row[10].number_format = 'general'
            row[11].value = row_result[1]
            row[11].number_format = 'general'
            row[11].alignment = Alignment(wrapText=True)
            row[12].value = row_result[2]
            row[12].number_format = '0%'
            row[13].value = row_result[3]
            row[13].number_format = 'general'
            row[14].value = row_result[4]
            row[14].number_format = 'general'
            row[14].alignment = Alignment(wrapText=True)
            row[15].value = row_result[5]
            row[15].number_format = '0%'
            #已重跑
            row[17].value = 1
        workbook.save(self.path)

    def copy_cell_value(self):
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook.worksheets[0]

        cell_value = sheet.cell(row=484, column=11).value
        #print(f'cell_value:{cell_value}')
        sheet.cell(row=485, column=11).value = cell_value
        sheet.cell(row=485, column=11).number_format = 'general'
        print(sheet.cell(row=484, column=12).value)
        sheet.cell(row=485, column=12).value = sheet.cell(row=484, column=12).value
        sheet.cell(row=485, column=12).number_format = '@'
        print(sheet.cell(row=484, column=13).value)
        sheet.cell(row=485, column=13).value = sheet.cell(row=484, column=13).value
        sheet.cell(row=485, column=13).number_format = '0%'
        workbook.save(self.path)

    def get_all_id_value(self):
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook.worksheets[0]
        id_list = []
        top1_list = []
        rows = sheet.rows
        for row in rows:
            id = row[0].value
            top1 = row[15].value
            if top1 is not None and top1 != '':
                id_list.append(id)
                top1_list.append(top1)
        print(len(id_list))
        print(id_list)
        print(top1_list)
        return id_list


    def write_path_length(self):
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook.worksheets[0]
        rows = list(sheet.rows)
        for row_index in range(1, len(rows)):
            path = rows[row_index][5].value
            print(f'path:{path}')
            path_l = ast.literal_eval(path)
            length = len(path_l)
            print(f' len: {length}')
            rows[row_index][14].value = length
        workbook.save(self.path)


if __name__ == '__main__':
    #resolver = Xls_resolve("/Users/liunian/Downloads/personal/论文相关/医疗实验/所有组实验结果.xlsx")
    #column_list = resolver.read_column_data(1)
    #print(column_list)
    #resolver.test()

    resolver = Xls_resolve("/Users/liunian/Downloads/personal/论文相关/医疗实验/另一组实验/another_factscore_0214.xlsx")
    #resolver.copy_cell_value()
    #result_list = resolver.get_left_result([483])
    #resolver.write_right_result([486], result_list)
    resolver.write_path_length()
    #resolver.get_all_id_value()